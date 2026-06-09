import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

import os

# ── colour palette ──────────────────────────────────────────────────────────
C_HEADER_DARK  = "1F4E79"   # deep navy
C_HEADER_MID   = "2E75B6"   # mid blue
C_HEADER_LIGHT = "BDD7EE"   # light blue
C_ACCENT       = "F4B942"   # amber
C_GREEN        = "70AD47"   # result green
C_GREEN_LIGHT  = "E2EFDA"
C_YELLOW_LIGHT = "FFF2CC"
C_ROW_ALT      = "DEEAF1"   # alternating row
C_WHITE        = "FFFFFF"
C_BORDER       = "2E75B6"

def border(style="thin"):
    s = Side(style=style, color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def hdr_font(size=11, bold=True, color=C_WHITE):
    return Font(name="Arial", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── 75 alternatif laptop ─────────────────────────────────────────────────────
# Kolom: (No, Nama Produk, Merek, TKDN(%), RAM(GB), SSD(GB), Garansi(thn), Harga(Rp))
alternatives = [
    (1,  "Lenovo IdeaPad Slim 3 15IAH8",      "Lenovo",  36.5, 8,  256, 1, 7800000),
    (2,  "Lenovo IdeaPad Slim 3 15IAN8",      "Lenovo",  36.5, 8,  512, 1, 8200000),
    (3,  "Lenovo ThinkPad E14 Gen 5",          "Lenovo",  38.2, 16, 512, 1, 14500000),
    (4,  "Lenovo ThinkPad L14 Gen 4",          "Lenovo",  38.2, 16, 512, 2, 16800000),
    (5,  "Lenovo ThinkBook 14 G6 IRL",         "Lenovo",  37.8, 16, 512, 1, 13200000),
    (6,  "ASUS Vivobook Go 14 E1404FA",        "ASUS",    40.1, 8,  512, 1, 7500000),
    (7,  "ASUS Vivobook 14 A1404ZA",           "ASUS",    40.1, 8,  512, 1, 8100000),
    (8,  "ASUS Vivobook 15 X1504ZA",           "ASUS",    40.1, 8,  256, 1, 7600000),
    (9,  "ASUS ExpertBook B1 B1402CVA",        "ASUS",    41.3, 16, 512, 1, 11900000),
    (10, "ASUS ExpertBook B1 B1502CVA",        "ASUS",    41.3, 16, 512, 1, 12300000),
    (11, "ASUS ExpertBook B5 B5402CVA",        "ASUS",    41.3, 16, 512, 2, 18700000),
    (12, "ASUS ExpertBook P1 P1403CVA",        "ASUS",    41.3, 8,  256, 1, 9800000),
    (13, "HP 14s-dq5002TU",                   "HP",      36.8, 8,  256, 1, 7300000),
    (14, "HP 14s-fq1019AU",                   "HP",      36.8, 8,  512, 1, 8000000),
    (15, "HP 15s-fq5012TU",                   "HP",      36.8, 8,  512, 1, 8400000),
    (16, "HP ProBook 440 G10",                "HP",      37.5, 16, 512, 1, 13500000),
    (17, "HP ProBook 450 G10",                "HP",      37.5, 16, 512, 1, 14100000),
    (18, "HP EliteBook 840 G10",              "HP",      38.9, 16, 512, 3, 22000000),
    (19, "HP EliteBook 640 G10",              "HP",      38.9, 16, 512, 3, 19500000),
    (20, "HP 250 G9",                         "HP",      36.8, 8,  256, 1, 7100000),
    (21, "Dell Inspiron 14 3420",             "Dell",    35.5, 8,  512, 1, 8300000),
    (22, "Dell Inspiron 15 3520",             "Dell",    35.5, 8,  512, 1, 8700000),
    (23, "Dell Vostro 3430",                  "Dell",    36.1, 8,  256, 1, 9200000),
    (24, "Dell Vostro 3520",                  "Dell",    36.1, 8,  512, 1, 9800000),
    (25, "Dell Vostro 5630",                  "Dell",    36.1, 16, 512, 1, 13900000),
    (26, "Dell Latitude 3440",               "Dell",    37.2, 8,  256, 3, 12500000),
    (27, "Dell Latitude 5440",               "Dell",    37.2, 16, 512, 3, 18200000),
    (28, "Dell Latitude 7440",               "Dell",    37.2, 16, 512, 3, 24500000),
    (29, "Acer Aspire 3 A314-36P",           "Acer",    36.2, 8,  256, 1, 7200000),
    (30, "Acer Aspire 5 A514-56P",           "Acer",    36.2, 8,  512, 1, 8500000),
    (31, "Acer Aspire 5 A515-58P",           "Acer",    36.2, 8,  512, 1, 8900000),
    (32, "Acer TravelMate P2 TMP214-54",     "Acer",    37.0, 8,  256, 1, 10500000),
    (33, "Acer TravelMate P4 TMP414-51",     "Acer",    37.0, 16, 512, 1, 14200000),
    (34, "Acer TravelMate P6 TMP614-52",     "Acer",    37.0, 16, 512, 3, 20100000),
    (35, "Axioo Mybook 14E",                 "Axioo",   40.5, 8,  256, 1, 6800000),
    (36, "Axioo Mybook 14H",                 "Axioo",   40.5, 8,  512, 1, 7400000),
    (37, "Axioo Mybook Pro M",               "Axioo",   40.5, 16, 512, 1, 10200000),
    (38, "Advan Workplus",                   "Advan",   40.8, 8,  256, 1, 6500000),
    (39, "Advan WorkPro 14",                 "Advan",   40.8, 8,  512, 1, 7000000),
    (40, "Advan WorkPro 15",                 "Advan",   40.8, 8,  512, 1, 7200000),
    (41, "Advan Jakarta i5",                 "Advan",   41.2, 16, 512, 1, 9500000),
    (42, "Zyrex Sky 232",                    "Zyrex",   40.3, 8,  256, 1, 6600000),
    (43, "Zyrex Sky 432",                    "Zyrex",   40.3, 8,  512, 1, 7100000),
    (44, "Zyrex Sky 532",                    "Zyrex",   40.3, 16, 512, 1, 9300000),
    (45, "MSI Modern 14 C13M",               "MSI",     35.8, 8,  512, 1, 9500000),
    (46, "MSI Modern 15 B12M",               "MSI",     35.8, 8,  512, 1, 9800000),
    (47, "MSI Prestige 14 Evo B13M",         "MSI",     35.8, 16, 512, 1, 14800000),
    (48, "ASUS Vivobook 16X K3605ZV",        "ASUS",    40.1, 16, 512, 1, 13700000),
    (49, "ASUS Zenbook 14 UX3402ZA",         "ASUS",    40.1, 16, 512, 1, 16200000),
    (50, "Lenovo IdeaPad 5 Pro 14IRH8",      "Lenovo",  36.5, 16, 512, 1, 14900000),
    (51, "HP ENVY x360 14-fc0xxx",           "HP",      37.5, 16, 512, 1, 17800000),
    (52, "Dell XPS 13 9340",                 "Dell",    35.5, 16, 512, 1, 24000000),
    (53, "Lenovo ThinkPad X1 Carbon Gen 11", "Lenovo",  38.2, 16, 512, 3, 32500000),
    (54, "HP ZBook Firefly 14 G9",           "HP",      38.9, 16, 512, 3, 28000000),
    (55, "Dell Latitude 9440",               "Dell",    37.2, 16, 512, 3, 30000000),
    (56, "Axioo Mybook Plus i7",             "Axioo",   40.5, 16, 512, 1, 11800000),
    (57, "ASUS ExpertBook B9 B9403CVA",      "ASUS",    41.3, 16, 512, 2, 23500000),
    (58, "Lenovo ThinkPad T14s Gen 4",       "Lenovo",  38.2, 16, 512, 3, 21000000),
    (59, "HP ProBook 650 G10",               "HP",      37.5, 16, 512, 3, 21500000),
    (60, "Dell Vostro 5530",                 "Dell",    36.1, 16, 512, 1, 14500000),
    (61, "Acer Swift 3 SF314-512",           "Acer",    36.2, 8,  512, 1, 10800000),
    (62, "Acer Swift Go 14 SFG14-71",        "Acer",    36.2, 16, 512, 1, 14500000),
    (63, "ASUS Vivobook 14X K3405ZF",        "ASUS",    40.1, 16, 512, 1, 12800000),
    (64, "Lenovo IdeaPad Gaming 3 15ARH7",   "Lenovo",  36.5, 16, 512, 1, 12500000),
    (65, "HP Pavilion 15-eg2xxx",            "HP",      36.8, 16, 512, 1, 12100000),
    (66, "Dell Inspiron 14 5440",            "Dell",    35.5, 16, 512, 1, 13300000),
    (67, "Advan Soulmate i5 1235U",          "Advan",   41.2, 16, 512, 1, 10500000),
    (68, "Zyrex Executive 1",                "Zyrex",   40.3, 16, 512, 2, 12000000),
    (69, "Axioo Mybook 14S Pro",             "Axioo",   40.5, 16, 512, 1, 11200000),
    (70, "ASUS BR1100CKA",                   "ASUS",    41.3, 8,  64,  1, 5900000),
    (71, "Lenovo 100e Chromebook Gen 3",     "Lenovo",  36.5, 8,  64,  1, 5500000),
    (72, "HP Chromebook 14a-nd0xxx",         "HP",      36.8, 8,  64,  1, 5700000),
    (73, "ASUS ExpertBook B1 B1400CEAE",     "ASUS",    41.3, 8,  256, 1, 10200000),
    (74, "Lenovo ThinkBook 15 G4 ABA",       "Lenovo",  37.8, 16, 512, 1, 12700000),
    (75, "HP ProBook 440 G9",                "HP",      37.5, 16, 512, 1, 12900000),
]

# ── Bobot & Jenis Kriteria ───────────────────────────────────────────────────
kriteria = [
    ("C1", "Nilai TKDN + BMP (%)", "Benefit", 0.30),
    ("C2", "Kapasitas RAM (GB)",  "Benefit", 0.25),
    ("C3", "Kapasitas SSD (GB)",  "Benefit", 0.20),
    ("C4", "Masa Garansi (Thn)",  "Benefit", 0.15),
    ("C5", "Harga Satuan (Rp)",   "Cost",    0.10),
]
weights = [k[3] for k in kriteria]

# ─────────────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 1 – COVER / PENDAHULUAN                                          ║
# ╚══════════════════════════════════════════════════════════════════════════╝
ws_cover = wb.active
ws_cover.title = "COVER"
ws_cover.sheet_view.showGridLines = False

ws_cover.column_dimensions["A"].width = 4
ws_cover.column_dimensions["B"].width = 50
ws_cover.column_dimensions["C"].width = 30
ws_cover.column_dimensions["D"].width = 20

# Background header
for r in range(1, 6):
    for c in range(1, 10):
        ws_cover.cell(r, c).fill = fill(C_HEADER_DARK)

ws_cover.merge_cells("B2:H4")
cell = ws_cover["B2"]
cell.value = "SISTEM PENDUKUNG KEPUTUSAN PENGADAAN PERANGKAT LAPTOP\nBERBASIS E-KATALOG MENGGUNAKAN METODE\nSIMPLE ADDITIVE WEIGHTING (SAW)\nPADA KEMENTERIAN PERDAGANGAN"
cell.font  = Font(name="Arial", size=16, bold=True, color=C_WHITE)
cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws_cover.row_dimensions[2].height = 100

info = [
    (8,  "Referensi Regulasi", "Kepmendag No. 2060 Tahun 2025"),
    (9,  "Metode",             "Simple Additive Weighting (SAW)"),
    (10, "Jumlah Alternatif",  "75 Laptop dari E-Katalog LKPP"),
    (11, "Jumlah Kriteria",    "5 Kriteria Evaluasi"),
    (12, "Sumber Data",        "katalog.inaproc.id / e-Katalog LKPP"),
    (14, "Kriteria 1 (C1)",    "Nilai TKDN (%) — Benefit — Bobot 30%"),
    (15, "Kriteria 2 (C2)",    "Kapasitas RAM (GB) — Benefit — Bobot 25%"),
    (16, "Kriteria 3 (C3)",    "Kapasitas SSD (GB) — Benefit — Bobot 20%"),
    (17, "Kriteria 4 (C4)",    "Masa Garansi (Tahun) — Benefit — Bobot 15%"),
    (18, "Kriteria 5 (C5)",    "Harga Satuan (Rp) — Cost — Bobot 10%"),
]
for row, label, val in info:
    c_label = ws_cover.cell(row, 2, label)
    c_label.font = Font(name="Arial", size=11, bold=True, color=C_HEADER_DARK)
    c_label.alignment = left()
    c_val = ws_cover.cell(row, 3, val)
    c_val.font = Font(name="Arial", size=11, color="000000")
    c_val.alignment = left()

ws_cover["B7"].value = "INFORMASI PENELITIAN"
ws_cover["B7"].font  = Font(name="Arial", size=12, bold=True, color=C_WHITE)
ws_cover["B7"].fill  = fill(C_HEADER_MID)
ws_cover["B7"].alignment = center()
ws_cover.merge_cells("B7:D7")

# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 2 – DATA ALTERNATIF                                              ║
# ╚══════════════════════════════════════════════════════════════════════════╝
ws_data = wb.create_sheet("DATA ALTERNATIF")
ws_data.sheet_view.showGridLines = False

# Column widths
col_widths = {"A":6,"B":8,"C":45,"D":15,"E":12,"F":12,"G":12,"H":12,"I":20}
for col, w in col_widths.items():
    ws_data.column_dimensions[col].width = w

# Title
ws_data.merge_cells("A1:I1")
t = ws_data["A1"]
t.value = "DATA ALTERNATIF LAPTOP – E-KATALOG LKPP"
t.font  = Font(name="Arial", size=14, bold=True, color=C_WHITE)
t.fill  = fill(C_HEADER_DARK)
t.alignment = center()
ws_data.row_dimensions[1].height = 30

ws_data.merge_cells("A2:I2")
sub = ws_data["A2"]
sub.value = "Referensi Spesifikasi Minimum: Kepmendag No. 2060 Tahun 2025"
sub.font  = Font(name="Arial", size=10, italic=True, color=C_HEADER_DARK)
sub.alignment = center()

# Header row
headers = ["No","Kode","Nama Produk / Model","Merek","C1\nTKDN (%)","C2\nRAM (GB)","C3\nSSD (GB)","C4\nGaransi (Thn)","C5\nHarga (Rp)"]
for col, h in enumerate(headers, 1):
    cell = ws_data.cell(4, col, h)
    cell.font  = hdr_font(10)
    cell.fill  = fill(C_HEADER_MID)
    cell.alignment = center()
    cell.border = border()
ws_data.row_dimensions[4].height = 40

# Data rows
for idx, alt in enumerate(alternatives):
    row = 5 + idx
    row_fill = fill(C_ROW_ALT) if idx % 2 == 0 else fill(C_WHITE)
    no, name, brand, tkdn, ram, ssd, warranty, price = alt
    kode = f"A{no:02d}"
    row_data = [no, kode, name, brand, tkdn, ram, ssd, warranty, price]
    for col, val in enumerate(row_data, 1):
        cell = ws_data.cell(row, col, val)
        cell.font   = body_font()
        cell.fill   = row_fill
        cell.border = border("thin")
        if col in (1, 2, 5, 6, 7, 8):
            cell.alignment = center()
        elif col == 9:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '#,##0'
        else:
            cell.alignment = left()

# Min/Max row at bottom for reference
ref_row = 5 + len(alternatives)
ws_data.merge_cells(f"A{ref_row}:D{ref_row}")
mr = ws_data[f"A{ref_row}"]
mr.value = "Min / Max Setiap Kriteria (referensi normalisasi)"
mr.font  = Font(name="Arial", size=10, bold=True, color=C_HEADER_DARK)
mr.alignment = center()
mr.fill = fill(C_YELLOW_LIGHT)

for col_idx, label in [(5,"MAX"),(6,"MAX"),(7,"MAX"),(8,"MAX"),(9,"MIN")]:
    col_letter = get_column_letter(col_idx)
    start_row  = 5
    end_row    = 5 + len(alternatives) - 1
    cell = ws_data.cell(ref_row, col_idx)
    if label == "MAX":
        cell.value = f"=MAX({col_letter}{start_row}:{col_letter}{end_row})"
    else:
        cell.value = f"=MIN({col_letter}{start_row}:{col_letter}{end_row})"
    cell.font   = Font(name="Arial", size=10, bold=True, color="000000")
    cell.fill   = fill(C_YELLOW_LIGHT)
    cell.border = border()
    cell.alignment = center()
    if col_idx == 9:
        cell.number_format = '#,##0'

# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 3 – BOBOT & KRITERIA (+ TABEL FUZZY)                             ║
# ╚══════════════════════════════════════════════════════════════════════════╝
ws_bobot = wb.create_sheet("BOBOT & KRITERIA")
ws_bobot.sheet_view.showGridLines = False

for col, w in {"A":5,"B":6,"C":34,"D":16,"E":5,"F":6,"G":34,"H":16}.items():
    ws_bobot.column_dimensions[col].width = w

# ── Title
ws_bobot.merge_cells("A1:H1")
t = ws_bobot["A1"]
t.value = "KRITERIA, BOBOT, DAN TABEL KEANGGOTAAN FUZZY – METODE SAW"
t.font  = Font(name="Arial", size=14, bold=True, color=C_WHITE)
t.fill  = fill(C_HEADER_DARK)
t.alignment = center()
ws_bobot.row_dimensions[1].height = 32

# ── Bagian 1: Ringkasan Kriteria & Bobot
ws_bobot.merge_cells("A3:H3")
s1 = ws_bobot["A3"]
s1.value = "BAGIAN 1 – RINGKASAN KRITERIA DAN BOBOT PENILAIAN"
s1.font  = Font(name="Arial", size=11, bold=True, color=C_WHITE)
s1.fill  = fill(C_HEADER_MID)
s1.alignment = center()
ws_bobot.row_dimensions[3].height = 22

bh = ["No","Kode","Nama Kriteria","Jenis","Bobot","Bobot (%)","Dasar Referensi",""]
for col, h in enumerate(bh[:7], 1):
    cell = ws_bobot.cell(4, col, h)
    cell.font  = hdr_font(10)
    cell.fill  = fill(C_HEADER_MID)
    cell.alignment = center()
    cell.border = border()
ws_bobot.merge_cells("G4:H4")
ws_bobot.row_dimensions[4].height = 30

basis = [
    "Kepmendag 2060/2025 – Min. TKDN 25%, BMP 40%",
    "Kepmendag 2060/2025 – Spek Min RAM: 8 GB",
    "Kepmendag 2060/2025 – Spek Min SSD: 256 GB",
    "Kepmendag 2060/2025 – Garansi minimal 1 tahun",
    "Efisiensi anggaran TIK (Diktum KETIGA & KEEMPAT)",
]
for i, (kode, nama, jenis, bobot) in enumerate(kriteria):
    row = 5 + i
    rf = fill(C_ROW_ALT) if i % 2 == 0 else fill(C_WHITE)
    for col, val in enumerate([i+1, kode, nama, jenis, bobot, bobot*100, basis[i]], 1):
        cell = ws_bobot.cell(row, col, val)
        cell.font   = body_font()
        cell.fill   = rf
        cell.border = border("thin")
        cell.alignment = center() if col in (1,2,4,5,6) else left()
        if col == 5: cell.number_format = "0.00"
        if col == 6: cell.number_format = '0"%"'
    ws_bobot.merge_cells(f"G{row}:H{row}")
    ws_bobot.cell(row,7).border = border("thin")

total_row_b = 5 + len(kriteria)
ws_bobot.merge_cells(f"A{total_row_b}:D{total_row_b}")
tc = ws_bobot[f"A{total_row_b}"]
tc.value = "TOTAL BOBOT"
tc.font  = Font(name="Arial", size=11, bold=True, color=C_WHITE)
tc.fill  = fill(C_HEADER_DARK)
tc.alignment = center(); tc.border = border()

for col_idx, val, fmt in [(5, sum(k[3] for k in kriteria), "0.00"),
                           (6, sum(k[3] for k in kriteria)*100, '0"%"')]:
    tc2 = ws_bobot.cell(total_row_b, col_idx, val)
    tc2.font  = Font(name="Arial", size=11, bold=True, color=C_WHITE)
    tc2.fill  = fill(C_GREEN); tc2.alignment = center()
    tc2.border = border(); tc2.number_format = fmt
ws_bobot.merge_cells(f"G{total_row_b}:H{total_row_b}")

# SAW formula notes
note_r = total_row_b + 2
ws_bobot.merge_cells(f"A{note_r}:H{note_r}")
note_t = ws_bobot[f"A{note_r}"]
note_t.value = "RUMUS SAW: Benefit → r_ij = f_ij / MAX(f_ij)   |   Cost → r_ij = MIN(f_ij) / f_ij   |   V_i = Σ (W_j × r_ij)"
note_t.font  = Font(name="Arial", size=10, bold=True, color=C_HEADER_DARK)
note_t.fill  = fill(C_YELLOW_LIGHT)
note_t.alignment = center()
note_t.border = border()
ws_bobot.row_dimensions[note_r].height = 20

# ── Bagian 2: Tabel Keanggotaan Fuzzy ────────────────────────────────────
fuz_start = note_r + 2
ws_bobot.merge_cells(f"A{fuz_start}:H{fuz_start}")
s2 = ws_bobot[f"A{fuz_start}"]
s2.value = "BAGIAN 2 – TABEL KEANGGOTAAN FUZZY SETIAP KRITERIA"
s2.font  = Font(name="Arial", size=11, bold=True, color=C_WHITE)
s2.fill  = fill(C_HEADER_MID)
s2.alignment = center()
ws_bobot.row_dimensions[fuz_start].height = 22

# Helper: gambar satu tabel fuzzy mulai di (start_row, col_start)
# col_start: 1=A,2=B,... tabel lebar 3 col: [No | Kondisi | Nilai]
def draw_fuzzy(ws, start_row, col_start, crit_type, sub_label, rows):
    hc = C_GREEN if crit_type == "Benefit" else C_ACCENT
    # baris 1: banner Benefit/Cost
    ws.merge_cells(start_row=start_row, start_column=col_start,
                   end_row=start_row, end_column=col_start+2)
    bn = ws.cell(start_row, col_start, crit_type.upper())
    bn.font = Font(name="Arial", size=10, bold=True, color=C_WHITE)
    bn.fill = fill(hc); bn.alignment = center(); bn.border = border()
    ws.row_dimensions[start_row].height = 18
    # baris 2: sub-header
    for c_off, txt in enumerate(["No", sub_label, "Nilai Fuzzy"]):
        sh = ws.cell(start_row+1, col_start+c_off, txt)
        sh.font = Font(name="Arial", size=9, bold=True, color=C_WHITE)
        sh.fill = fill(C_HEADER_MID); sh.alignment = center(); sh.border = border()
    ws.row_dimensions[start_row+1].height = 26
    # baris data
    for i, (cond, val) in enumerate(rows):
        r = start_row + 2 + i
        rf = fill(C_ROW_ALT) if i % 2 == 0 else fill(C_WHITE)
        no_c = ws.cell(r, col_start, i+1)
        no_c.font = Font(name="Arial", size=9, bold=True, color=C_WHITE)
        no_c.fill = fill(C_HEADER_MID); no_c.alignment = center(); no_c.border = border("thin")
        cd = ws.cell(r, col_start+1, cond)
        cd.font = body_font(9); cd.fill = rf; cd.alignment = left(); cd.border = border("thin")
        vl = ws.cell(r, col_start+2, val)
        vl.font = Font(name="Arial", size=9, bold=True)
        val_fc = C_GREEN_LIGHT if crit_type == "Benefit" else C_YELLOW_LIGHT
        vl.fill = fill(val_fc); vl.alignment = center(); vl.border = border("thin")
        vl.number_format = "0.00"
        ws.row_dimensions[r].height = 16

# Definisi tabel fuzzy per kriteria
fuzzy_tables = [
    ("Benefit", "Nilai TKDN + BMP (%)", [
        ("TKDN < 45%",               0.25),
        ("45% ≤ TKDN < 50%",         0.50),
        ("50% ≤ TKDN < 55%",         0.75),
        ("TKDN ≥ 55%",               1.00),
    ]),
    ("Benefit", "Kapasitas RAM (GB)", [
        ("RAM = 8 GB",                0.25),
        ("RAM = 16 GB",               0.50),
        ("RAM = 32 GB",               0.75),
        ("RAM > 32 GB",               1.00),
    ]),
    ("Benefit", "Kapasitas SSD (GB)", [
        ("SSD = 256 GB",              0.33),
        ("SSD = 512 GB",              0.67),
        ("SSD ≥ 1024 GB (1 TB)",      1.00),
    ]),
    ("Benefit", "Masa Garansi (Tahun)", [
        ("Garansi = 1 Tahun",         0.25),
        ("Garansi = 2 Tahun",         0.50),
        ("Garansi = 3 Tahun",         0.75),
        ("Garansi > 3 Tahun",         1.00),
    ]),
    ("Cost", "Harga Satuan (Rp)", [
        ("Harga ≤ Rp 11.000.000",     0.25),
        ("Rp 11 jt < Harga ≤ Rp 14 jt", 0.50),
        ("Rp 14 jt < Harga ≤ Rp 17 jt", 0.75),
        ("Harga > Rp 17.000.000",     1.00),
    ]),
]

# Layout: 2 kolom (col A=1, col E=5), max 3 tabel kiri / 2 kanan
# Posisi tabel: (start_row, col_start)
# Kiri (col A=1): C1, C3, C5 | Kanan (col E=5): C2, C4
left_col  = 1   # col A
right_col = 5   # col E
left_row  = fuz_start + 2
right_row = fuz_start + 2

for tbl_idx, (ctype, slabel, rows) in enumerate(fuzzy_tables):
    n_rows = len(rows)  # jumlah kategori
    tbl_height = n_rows + 2   # banner + subheader + data
    if tbl_idx % 2 == 0:  # kiri
        draw_fuzzy(ws_bobot, left_row, left_col, ctype, slabel, rows)
        left_row += tbl_height + 2  # gap 2 baris
    else:                  # kanan
        draw_fuzzy(ws_bobot, right_row, right_col, ctype, slabel, rows)
        right_row += tbl_height + 2

# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 4 – MATRIKS KEPUTUSAN                                            ║
# ╚══════════════════════════════════════════════════════════════════════════╝
ws_matrix = wb.create_sheet("MATRIKS KEPUTUSAN")
ws_matrix.sheet_view.showGridLines = False

for col, w in {"A":6,"B":8,"C":42,"D":14,"E":14,"F":14,"G":14,"H":22}.items():
    ws_matrix.column_dimensions[col].width = w

ws_matrix.merge_cells("A1:H1")
t = ws_matrix["A1"]
t.value = "MATRIKS KEPUTUSAN (X) – DATA MENTAH SETIAP ALTERNATIF DAN KRITERIA"
t.font  = Font(name="Arial", size=13, bold=True, color=C_WHITE)
t.fill  = fill(C_HEADER_DARK)
t.alignment = center()
ws_matrix.row_dimensions[1].height = 28

headers_mx = ["No","Kode","Nama Produk / Model","C1\nTKDN (%)","C2\nRAM (GB)","C3\nSSD (GB)","C4\nGaransi (Thn)","C5\nHarga (Rp)"]
for col, h in enumerate(headers_mx, 1):
    cell = ws_matrix.cell(3, col, h)
    cell.font  = hdr_font(10)
    cell.fill  = fill(C_HEADER_MID)
    cell.alignment = center()
    cell.border = border()
ws_matrix.row_dimensions[3].height = 40

# Reference from DATA ALTERNATIF sheet
for idx in range(len(alternatives)):
    row = 4 + idx
    alt_row = 5 + idx   # row in DATA ALTERNATIF
    row_fill = fill(C_ROW_ALT) if idx % 2 == 0 else fill(C_WHITE)

    # No, Kode, Name, Brand are copied/referenced from DATA sheet
    for col, src_col in [(1,"A"),(2,"B"),(3,"C")]:
        cell = ws_matrix.cell(row, col)
        cell.value = f"='DATA ALTERNATIF'!{src_col}{alt_row}"
        cell.font  = body_font()
        cell.fill  = row_fill
        cell.border = border("thin")
        cell.alignment = center() if col < 3 else left()

    # C1..C5 references
    for col, src_col in [(4,"E"),(5,"F"),(6,"G"),(7,"H"),(8,"I")]:
        cell = ws_matrix.cell(row, col)
        cell.value = f"='DATA ALTERNATIF'!{src_col}{alt_row}"
        cell.font  = body_font()
        cell.fill  = row_fill
        cell.border = border("thin")
        cell.alignment = center()
        if col == 8:
            cell.number_format = '#,##0'

# Max/Min row
ref_row_mx = 4 + len(alternatives)
ws_matrix.merge_cells(f"A{ref_row_mx}:C{ref_row_mx}")
mr = ws_matrix[f"A{ref_row_mx}"]
mr.value = "MAX (Benefit) / MIN (Cost)"
mr.font  = Font(name="Arial", size=10, bold=True, color=C_HEADER_DARK)
mr.fill  = fill(C_YELLOW_LIGHT)
mr.alignment = center()
mr.border = border()

for col_idx, label in [(4,"MAX"),(5,"MAX"),(6,"MAX"),(7,"MAX"),(8,"MIN")]:
    cl = get_column_letter(col_idx)
    cell = ws_matrix.cell(ref_row_mx, col_idx)
    if label == "MAX":
        cell.value = f"=MAX({cl}4:{cl}{ref_row_mx-1})"
    else:
        cell.value = f"=MIN({cl}4:{cl}{ref_row_mx-1})"
    cell.font   = Font(name="Arial", size=10, bold=True)
    cell.fill   = fill(C_YELLOW_LIGHT)
    cell.border = border()
    cell.alignment = center()
    if col_idx == 8:
        cell.number_format = '#,##0'

# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 5 – NILAI FUZZY                                                  ║
# ╚══════════════════════════════════════════════════════════════════════════╝
ws_fuzzy = wb.create_sheet("NILAI FUZZY")
ws_fuzzy.sheet_view.showGridLines = False

for col, w in {"A":6,"B":8,"C":42,"D":14,"E":14,"F":14,"G":14,"H":16}.items():
    ws_fuzzy.column_dimensions[col].width = w

ws_fuzzy.merge_cells("A1:H1")
tf = ws_fuzzy["A1"]
tf.value = "NILAI FUZZY (f_ij) – KONVERSI DATA MENTAH KE NILAI KEANGGOTAAN FUZZY"
tf.font  = Font(name="Arial", size=12, bold=True, color=C_WHITE)
tf.fill  = fill(C_HEADER_DARK)
tf.alignment = center()
ws_fuzzy.row_dimensions[1].height = 28

ws_fuzzy.merge_cells("A2:H2")
sub_f = ws_fuzzy["A2"]
sub_f.value = ("Benefit → Semakin tinggi nilai fuzzy semakin baik  |  "
               "Cost → Nilai fuzzy mencerminkan tingkat harga (normalisasi akan membalik)")
sub_f.font  = Font(name="Arial", size=9, italic=True, color=C_HEADER_DARK)
sub_f.alignment = center()

headers_fz = ["No","Kode","Nama Produk / Model",
              "f_C1\nTKDN+BMP","f_C2\nRAM","f_C3\nSSD","f_C4\nGaransi","f_C5\nHarga"]
for col, h in enumerate(headers_fz, 1):
    cell = ws_fuzzy.cell(4, col, h)
    cell.font  = hdr_font(10)
    cell.fill  = fill(C_HEADER_MID)
    cell.alignment = center()
    cell.border = border()
ws_fuzzy.row_dimensions[4].height = 40

# Rumus fuzzy IF per kriteria, referensi dari MATRIKS KEPUTUSAN
# mx_col: kolom di sheet MATRIKS KEPUTUSAN (D=C1, E=C2, F=C3, G=C4, H=C5)
fuzzy_formulas = {
    # col idx in ws_fuzzy → (MK_col, IF formula template)
    4: lambda r: (f"=IF('MATRIKS KEPUTUSAN'!D{r}<45,0.25,"
                  f"IF('MATRIKS KEPUTUSAN'!D{r}<50,0.5,"
                  f"IF('MATRIKS KEPUTUSAN'!D{r}<55,0.75,1)))"),
    5: lambda r: (f"=IF('MATRIKS KEPUTUSAN'!E{r}<=8,0.25,"
                  f"IF('MATRIKS KEPUTUSAN'!E{r}<=16,0.5,"
                  f"IF('MATRIKS KEPUTUSAN'!E{r}<=32,0.75,1)))"),
    6: lambda r: (f"=IF('MATRIKS KEPUTUSAN'!F{r}<=256,0.33,"
                  f"IF('MATRIKS KEPUTUSAN'!F{r}<=512,0.67,1))"),
    7: lambda r: (f"=IF('MATRIKS KEPUTUSAN'!G{r}<=1,0.25,"
                  f"IF('MATRIKS KEPUTUSAN'!G{r}<=2,0.5,"
                  f"IF('MATRIKS KEPUTUSAN'!G{r}<=3,0.75,1)))"),
    8: lambda r: (f"=IF('MATRIKS KEPUTUSAN'!H{r}<=11000000,0.25,"
                  f"IF('MATRIKS KEPUTUSAN'!H{r}<=14000000,0.5,"
                  f"IF('MATRIKS KEPUTUSAN'!H{r}<=17000000,0.75,1)))"),
}

for idx in range(len(alternatives)):
    row    = 4 + idx
    mx_row = 4 + idx
    row_fill = fill(C_ROW_ALT) if idx % 2 == 0 else fill(C_WHITE)

    for col, src_col in [(1,"A"),(2,"B"),(3,"C")]:
        cell = ws_fuzzy.cell(row, col)
        cell.value = f"='MATRIKS KEPUTUSAN'!{src_col}{mx_row}"
        cell.font  = body_font(); cell.fill  = row_fill
        cell.border = border("thin")
        cell.alignment = center() if col < 3 else left()

    for col_idx, formula_fn in fuzzy_formulas.items():
        cell = ws_fuzzy.cell(row, col_idx)
        cell.value  = formula_fn(mx_row)
        cell.font   = body_font()
        cell.fill   = row_fill
        cell.border = border("thin")
        cell.alignment = center()
        cell.number_format = "0.00"

# MAX / MIN row di bawah sebagai referensi normalisasi
fz_ref_row = 4 + len(alternatives)
ws_fuzzy.merge_cells(f"A{fz_ref_row}:C{fz_ref_row}")
mr_fz = ws_fuzzy[f"A{fz_ref_row}"]
mr_fz.value = "MAX Benefit / MIN Cost (referensi normalisasi)"
mr_fz.font  = Font(name="Arial", size=9, bold=True, color=C_HEADER_DARK)
mr_fz.fill  = fill(C_YELLOW_LIGHT); mr_fz.alignment = center(); mr_fz.border = border()

for col_idx, label in [(4,"MAX"),(5,"MAX"),(6,"MAX"),(7,"MAX"),(8,"MIN")]:
    cl = get_column_letter(col_idx)
    cell = ws_fuzzy.cell(fz_ref_row, col_idx)
    cell.value = f"={'MAX' if label=='MAX' else 'MIN'}({cl}4:{cl}{fz_ref_row-1})"
    cell.font  = Font(name="Arial", size=9, bold=True)
    cell.fill  = fill(C_YELLOW_LIGHT); cell.border = border(); cell.alignment = center()
    cell.number_format = "0.00"

# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 6 – NORMALISASI (menggunakan nilai fuzzy)                        ║
# ╚══════════════════════════════════════════════════════════════════════════╝
ws_norm = wb.create_sheet("NORMALISASI")
ws_norm.sheet_view.showGridLines = False

for col, w in {"A":6,"B":8,"C":42,"D":14,"E":14,"F":14,"G":14,"H":16}.items():
    ws_norm.column_dimensions[col].width = w

ws_norm.merge_cells("A1:H1")
t = ws_norm["A1"]
t.value = "MATRIKS NORMALISASI (R) – Benefit: r_ij = f_ij / 1   |   Cost: r_ij = 0,25 / f_ij"
t.font  = Font(name="Arial", size=12, bold=True, color=C_WHITE)
t.fill  = fill(C_HEADER_DARK)
t.alignment = center()
ws_norm.row_dimensions[1].height = 28

ws_norm.merge_cells("A2:H2")
sub_n = ws_norm["A2"]
sub_n.value = ("MAX nilai fuzzy Benefit = 1,00  →  r_ij = f_ij / 1 = f_ij     |     "
               "MIN nilai fuzzy Cost = 0,25  →  r_ij = 0,25 / f_ij")
sub_n.font  = Font(name="Arial", size=9, italic=True, color=C_HEADER_DARK)
sub_n.alignment = center()

headers_nrm = ["No","Kode","Nama Produk / Model",
               "r_C1\nTKDN+BMP","r_C2\nRAM","r_C3\nSSD","r_C4\nGaransi","r_C5\nHarga"]
for col, h in enumerate(headers_nrm, 1):
    cell = ws_norm.cell(3, col, h)
    cell.font  = hdr_font(10)
    cell.fill  = fill(C_HEADER_MID)
    cell.alignment = center()
    cell.border = border()
ws_norm.row_dimensions[3].height = 40

for idx in range(len(alternatives)):
    row      = 4 + idx
    fz_row   = 4 + idx   # baris di sheet NILAI FUZZY
    row_fill = fill(C_ROW_ALT) if idx % 2 == 0 else fill(C_WHITE)

    for col, src_col in [(1,"A"),(2,"B"),(3,"C")]:
        cell = ws_norm.cell(row, col)
        cell.value = f"='NILAI FUZZY'!{src_col}{fz_row}"
        cell.font  = body_font(); cell.fill = row_fill
        cell.border = border("thin")
        cell.alignment = center() if col < 3 else left()

    # C1–C4 Benefit: r_ij = f_ij / MAX(f_j) = f_ij / 1 = f_ij
    for col_idx, fz_col in [(4,"D"),(5,"E"),(6,"F"),(7,"G")]:
        cell = ws_norm.cell(row, col_idx)
        cell.value = f"='NILAI FUZZY'!{fz_col}{fz_row}"
        cell.font  = body_font(); cell.fill = row_fill
        cell.border = border("thin"); cell.alignment = center()
        cell.number_format = "0.0000"

    # C5 Cost: r_ij = MIN(f_j) / f_ij = 0.25 / f_ij
    cell = ws_norm.cell(row, 8)
    cell.value = f"=0.25/'NILAI FUZZY'!H{fz_row}"
    cell.font  = body_font(); cell.fill = row_fill
    cell.border = border("thin"); cell.alignment = center()
    cell.number_format = "0.0000"

# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 6 – NILAI PREFERENSI & RANKING                                   ║
# ╚══════════════════════════════════════════════════════════════════════════╝
ws_rank = wb.create_sheet("NILAI PREFERENSI & RANKING")
ws_rank.sheet_view.showGridLines = False

for col, w in {"A":7,"B":8,"C":44,"D":14,"E":14,"F":14,"G":14,"H":14,"I":16,"J":10,"K":12}.items():
    ws_rank.column_dimensions[col].width = w

ws_rank.merge_cells("A1:K1")
t = ws_rank["A1"]
t.value = "NILAI PREFERENSI (V_i) DAN RANKING – V_i = Σ (W_j × r_ij)"
t.font  = Font(name="Arial", size=13, bold=True, color=C_WHITE)
t.fill  = fill(C_HEADER_DARK)
t.alignment = center()
ws_rank.row_dimensions[1].height = 28

# Bobot row
ws_rank.merge_cells("A2:C2")
bw_label = ws_rank["A2"]
bw_label.value = "Bobot (W_j):"
bw_label.font  = Font(name="Arial", size=10, bold=True, color=C_HEADER_DARK)
bw_label.alignment = center()
bw_label.fill  = fill(C_YELLOW_LIGHT)
bw_label.border = border()

bobot_vals = [0.30, 0.25, 0.20, 0.15, 0.10]
for col_idx, bw in zip([4,5,6,7,8], bobot_vals):
    cell = ws_rank.cell(2, col_idx, bw)
    cell.font   = Font(name="Arial", size=10, bold=True, color=C_HEADER_DARK)
    cell.fill   = fill(C_YELLOW_LIGHT)
    cell.border = border()
    cell.alignment = center()
    cell.number_format = "0.00"

headers_rk = ["No","Kode","Nama Produk / Model",
               "W×r_C1\nTKDN","W×r_C2\nRAM","W×r_C3\nSSD","W×r_C4\nGaransi","W×r_C5\nHarga",
               "V_i\n(Nilai Preferensi)","RANK","RANK\nUnik"]
for col, h in enumerate(headers_rk, 1):
    cell = ws_rank.cell(3, col, h)
    cell.font  = hdr_font(10)
    cell.fill  = fill(C_HEADER_MID)
    cell.alignment = center()
    cell.border = border()
ws_rank.row_dimensions[3].height = 45

for idx in range(len(alternatives)):
    row = 4 + idx
    nrm_row = 4 + idx  # row in NORMALISASI sheet
    row_fill = fill(C_ROW_ALT) if idx % 2 == 0 else fill(C_WHITE)

    for col, src_col in [(1,"A"),(2,"B"),(3,"C")]:
        cell = ws_rank.cell(row, col)
        cell.value = f"='NORMALISASI'!{src_col}{nrm_row}"
        cell.font  = body_font()
        cell.fill  = row_fill
        cell.border = border("thin")
        cell.alignment = center() if col < 3 else left()

    # W × r for each criterion
    bobot_row = 2  # row with bobot in this sheet
    norm_cols = ["D","E","F","G","H"]
    rank_cols_idx = [4,5,6,7,8]
    for col_idx, norm_col in zip(rank_cols_idx, norm_cols):
        bobot_col = get_column_letter(col_idx)
        cell = ws_rank.cell(row, col_idx)
        cell.value = f"={bobot_col}{bobot_row}*'NORMALISASI'!{norm_col}{nrm_row}"
        cell.font  = body_font()
        cell.fill  = row_fill
        cell.border = border("thin")
        cell.alignment = center()
        cell.number_format = "0.0000"

    # V_i = sum of W×r
    vi_cell = ws_rank.cell(row, 9)
    vi_cell.value = f"=SUM(D{row}:H{row})"
    vi_cell.font  = Font(name="Arial", size=10, bold=True, color="000000")
    vi_cell.fill  = row_fill
    vi_cell.border = border()
    vi_cell.alignment = center()
    vi_cell.number_format = "0.0000"

    # RANK (dengan seri)
    rank_cell = ws_rank.cell(row, 10)
    rank_cell.value = f"=RANK(I{row},I$4:I${4+len(alternatives)-1},0)"
    rank_cell.font  = Font(name="Arial", size=10, bold=True)
    rank_cell.fill  = row_fill
    rank_cell.border = border()
    rank_cell.alignment = center()

    # RANK UNIK – memecah seri dengan COUNTIF agar setiap baris dapat nomor unik
    # Rumus: RANK normal + jumlah kemunculan V_i yang sama di atas baris ini
    # Hasilnya: laptop seri pertama tetap di rank asli, berikutnya +1, +2, dst
    uniq_cell = ws_rank.cell(row, 11)
    uniq_cell.value = (f"=RANK(I{row},I$4:I${4+len(alternatives)-1},0)"
                       f"+COUNTIF($I$4:I{row},I{row})-1")
    uniq_cell.font  = Font(name="Arial", size=10, bold=True, color=C_HEADER_DARK)
    uniq_cell.fill  = fill(C_YELLOW_LIGHT)
    uniq_cell.border = border()
    uniq_cell.alignment = center()

# Conditional formatting hint – highlight top 10
# (openpyxl supports it but keep simple)

# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 7 – HASIL AKHIR (TOP 75 SORTED)                                  ║
# ╚══════════════════════════════════════════════════════════════════════════╝
ws_hasil = wb.create_sheet("HASIL AKHIR")
ws_hasil.sheet_view.showGridLines = False

for col, w in {"A":7,"B":8,"C":44,"D":16,"E":13,"F":13,"G":13,"H":13,"I":22,"J":18,"K":10}.items():
    ws_hasil.column_dimensions[col].width = w

ws_hasil.merge_cells("A1:K1")
t = ws_hasil["A1"]
t.value = "HASIL AKHIR – RANKING LAPTOP TERBAIK BERDASARKAN METODE SAW"
t.font  = Font(name="Arial", size=13, bold=True, color=C_WHITE)
t.fill  = fill(C_HEADER_DARK)
t.alignment = center()
ws_hasil.row_dimensions[1].height = 28

ws_hasil.merge_cells("A2:K2")
sub = ws_hasil["A2"]
sub.value = "Data diurutkan berdasarkan nilai preferensi V_i tertinggi → terendah (SMALL/LARGE formula menggunakan helper kolom)"
sub.font  = Font(name="Arial", size=9, italic=True, color=C_HEADER_DARK)
sub.alignment = center()

headers_h = ["RANK","Kode","Nama Produk / Model","Merek",
             "C1\nTKDN (%)","C2\nRAM (GB)","C3\nSSD (GB)","C4\nGaransi (Thn)",
             "C5\nHarga (Rp)","V_i\n(Nilai Preferensi)","KETERANGAN"]
for col, h in enumerate(headers_h, 1):
    cell = ws_hasil.cell(4, col, h)
    cell.font  = hdr_font(10)
    cell.fill  = fill(C_HEADER_MID)
    cell.alignment = center()
    cell.border = border()
ws_hasil.row_dimensions[4].height = 40

# HASIL AKHIR: gunakan kolom RANK UNIK (K) dari sheet NILAI PREFERENSI & RANKING
# sehingga setiap rank_num (1,2,3...) pasti menemukan tepat satu baris unik
# → tidak ada duplikat meskipun ada seri V_i
da_name  = "'DATA ALTERNATIF'"
rk_sheet = "'NILAI PREFERENSI & RANKING'"
uniq_rank_range = f"{rk_sheet}!$K$4:$K${4+len(alternatives)-1}"
vi_range        = f"{rk_sheet}!$I$4:$I${4+len(alternatives)-1}"

for rank_num in range(1, len(alternatives)+1):
    row = 4 + rank_num
    row_fill = fill(C_GREEN_LIGHT) if rank_num <= 10 else (fill(C_ROW_ALT) if rank_num % 2 == 0 else fill(C_WHITE))

    # RANK column
    rk = ws_hasil.cell(row, 1, rank_num)
    rk.font  = Font(name="Arial", size=10, bold=(rank_num<=10))
    rk.fill  = row_fill
    rk.border = border("thin")
    rk.alignment = center()

    # MATCH ke kolom RANK UNIK → selalu temukan tepat 1 baris
    match_formula = f"MATCH({rank_num},{uniq_rank_range},0)"

    # V_i: INDEX ke kolom I NILAI PREFERENSI berdasarkan posisi yang ditemukan MATCH
    vi_formula = f"=INDEX({vi_range},{match_formula})"

    kode_f = f"=INDEX({da_name}!$B$5:$B${5+len(alternatives)-1},{match_formula})"
    name_f = f"=INDEX({da_name}!$C$5:$C${5+len(alternatives)-1},{match_formula})"
    brand_f= f"=INDEX({da_name}!$D$5:$D${5+len(alternatives)-1},{match_formula})"
    c1_f   = f"=INDEX({da_name}!$E$5:$E${5+len(alternatives)-1},{match_formula})"
    c2_f   = f"=INDEX({da_name}!$F$5:$F${5+len(alternatives)-1},{match_formula})"
    c3_f   = f"=INDEX({da_name}!$G$5:$G${5+len(alternatives)-1},{match_formula})"
    c4_f   = f"=INDEX({da_name}!$H$5:$H${5+len(alternatives)-1},{match_formula})"
    c5_f   = f"=INDEX({da_name}!$I$5:$I${5+len(alternatives)-1},{match_formula})"

    cols_vals = [
        (2,  kode_f,  center(), None),
        (3,  name_f,  left(),   None),
        (4,  brand_f, center(), None),
        (5,  c1_f,    center(), "0.0"),
        (6,  c2_f,    center(), None),
        (7,  c3_f,    center(), None),
        (8,  c4_f,    center(), None),
        (9,  c5_f,    Alignment(horizontal="right", vertical="center"), '#,##0'),
        (10, vi_formula, center(), "0.0000"),
    ]
    for col_idx, formula, align, numfmt in cols_vals:
        cell = ws_hasil.cell(row, col_idx, formula)
        cell.font   = Font(name="Arial", size=10, bold=(rank_num<=10))
        cell.fill   = row_fill
        cell.border = border("thin")
        cell.alignment = align
        if numfmt:
            cell.number_format = numfmt

    # Keterangan
    ket_cell = ws_hasil.cell(row, 11)
    if rank_num <= 3:
        ket_cell.value = "⭐ SANGAT DIREKOMENDASIKAN"
        ket_cell.font  = Font(name="Arial", size=9, bold=True, color="000000")
        ket_cell.fill  = fill(C_GREEN)
    elif rank_num <= 10:
        ket_cell.value = "✔ DIREKOMENDASIKAN"
        ket_cell.font  = Font(name="Arial", size=9, color="000000")
        ket_cell.fill  = fill(C_GREEN_LIGHT)
    else:
        ket_cell.value = "Memenuhi syarat"
        ket_cell.font  = Font(name="Arial", size=9, color="595959")
        ket_cell.fill  = row_fill
    ket_cell.border = border("thin")
    ket_cell.alignment = center()

# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 8 – CONTOH PERHITUNGAN MANUAL (1 alternatif detail)              ║
# ╚══════════════════════════════════════════════════════════════════════════╝
ws_manual = wb.create_sheet("CONTOH PERHITUNGAN")
ws_manual.sheet_view.showGridLines = False
for col, w in {"A":5,"B":35,"C":22,"D":22,"E":22,"F":18,"G":18}.items():
    ws_manual.column_dimensions[col].width = w

ws_manual.merge_cells("A1:G1")
t = ws_manual["A1"]
t.value = "CONTOH LANGKAH PERHITUNGAN MANUAL SAW – ALTERNATIF A01 s/d A05"
t.font  = Font(name="Arial", size=13, bold=True, color=C_WHITE)
t.fill  = fill(C_HEADER_DARK)
t.alignment = center()
ws_manual.row_dimensions[1].height = 28

steps = [
    (3,  "LANGKAH 1: Tentukan Kriteria dan Bobot", C_HEADER_MID),
    (4,  "Kriteria", None),
    (5,  "C1 – Nilai TKDN (%)", None),
    (6,  "C2 – Kapasitas RAM (GB)", None),
    (7,  "C3 – Kapasitas SSD (GB)", None),
    (8,  "C4 – Masa Garansi (Tahun)", None),
    (9,  "C5 – Harga Satuan (Rp)", None),
]
step1_data = [
    (4, "Kriteria", "Jenis", "Bobot"),
    (5, "C1 – Nilai TKDN (%)",       "Benefit", "0.30"),
    (6, "C2 – Kapasitas RAM (GB)",   "Benefit", "0.25"),
    (7, "C3 – Kapasitas SSD (GB)",   "Benefit", "0.20"),
    (8, "C4 – Masa Garansi (Tahun)", "Benefit", "0.15"),
    (9, "C5 – Harga Satuan (Rp)",    "Cost",    "0.10"),
]
ws_manual.cell(3, 2, "LANGKAH 1: Tentukan Kriteria dan Bobot").font = Font(name="Arial", size=11, bold=True, color=C_WHITE)
ws_manual.cell(3, 2).fill = fill(C_HEADER_MID)
ws_manual.merge_cells("B3:G3")
ws_manual.cell(3, 2).alignment = center()

for row, a, b, c in step1_data:
    for col, val in [(2,a),(3,b),(4,c)]:
        cell = ws_manual.cell(row, col, val)
        cell.font   = Font(name="Arial", size=10, bold=(row==4))
        cell.fill   = fill(C_ROW_ALT) if row % 2 == 0 else fill(C_WHITE)
        cell.border = border("thin")
        cell.alignment = center()

ws_manual.merge_cells("B11:G11")
s2 = ws_manual.cell(11, 2, "LANGKAH 2: Buat Matriks Keputusan X (ambil 5 alternatif sebagai contoh)")
s2.font = Font(name="Arial", size=11, bold=True, color=C_WHITE)
s2.fill = fill(C_HEADER_MID)
s2.alignment = center()

sample_5 = alternatives[:5]
mx_hdr = [(12,"Alternatif","C1 TKDN","C2 RAM","C3 SSD","C4 Garansi","C5 Harga")]
for row, *vals in mx_hdr:
    for col, v in enumerate(["Alternatif","C1 TKDN","C2 RAM","C3 SSD","C4 Garansi","C5 Harga"], 2):
        cell = ws_manual.cell(row, col, v)
        cell.font  = hdr_font(10)
        cell.fill  = fill(C_HEADER_MID)
        cell.alignment = center()
        cell.border = border()

for i, alt in enumerate(sample_5):
    row = 13 + i
    no, name, brand, tkdn, ram, ssd, warranty, price = alt
    vals = [f"A{no:02d} – {name[:30]}", tkdn, ram, ssd, warranty, price]
    for col, v in enumerate(vals, 2):
        cell = ws_manual.cell(row, col, v)
        cell.font  = body_font()
        cell.fill  = fill(C_ROW_ALT) if i % 2 == 0 else fill(C_WHITE)
        cell.border = border("thin")
        cell.alignment = Alignment(horizontal="right" if col>2 else "left", vertical="center")
        if col == 7:
            cell.number_format = '#,##0'

ws_manual.merge_cells("B19:G19")
s3 = ws_manual.cell(19, 2, "LANGKAH 3: Normalisasi Matriks R  |  Benefit: r_ij = x_ij / MAX(x_ij)   |   Cost: r_ij = MIN(x_ij) / x_ij")
s3.font = Font(name="Arial", size=11, bold=True, color=C_WHITE)
s3.fill = fill(C_HEADER_MID)
s3.alignment = center()

# compute normalization manually for 5 samples
tkdn_vals  = [a[3] for a in sample_5]
ram_vals   = [a[4] for a in sample_5]
ssd_vals   = [a[5] for a in sample_5]
war_vals   = [a[6] for a in sample_5]
price_vals = [a[7] for a in sample_5]

max_tkdn  = max(tkdn_vals)
max_ram   = max(ram_vals)
max_ssd   = max(ssd_vals)
max_war   = max(war_vals)
min_price = min(price_vals)

norm_hdr = ["Alternatif","r_C1","r_C2","r_C3","r_C4","r_C5"]
for col, v in enumerate(norm_hdr, 2):
    cell = ws_manual.cell(20, col, v)
    cell.font  = hdr_font(10)
    cell.fill  = fill(C_HEADER_MID)
    cell.alignment = center()
    cell.border = border()

for i, alt in enumerate(sample_5):
    row = 21 + i
    no, name, brand, tkdn, ram, ssd, warranty, price = alt
    r1 = round(tkdn / max_tkdn, 4)
    r2 = round(ram  / max_ram,  4)
    r3 = round(ssd  / max_ssd,  4)
    r4 = round(warranty / max_war, 4)
    r5 = round(min_price / price,  4)
    vals = [f"A{no:02d}", r1, r2, r3, r4, r5]
    for col, v in enumerate(vals, 2):
        cell = ws_manual.cell(row, col, v)
        cell.font  = body_font()
        cell.fill  = fill(C_ROW_ALT) if i % 2 == 0 else fill(C_WHITE)
        cell.border = border("thin")
        cell.alignment = center()
        if col > 2:
            cell.number_format = "0.0000"

ws_manual.merge_cells("B27:G27")
s4 = ws_manual.cell(27, 2, "LANGKAH 4: Hitung Nilai Preferensi V_i = (0.30×r_C1) + (0.25×r_C2) + (0.20×r_C3) + (0.15×r_C4) + (0.10×r_C5)")
s4.font = Font(name="Arial", size=11, bold=True, color=C_WHITE)
s4.fill = fill(C_HEADER_MID)
s4.alignment = center()

vi_hdr = ["Alternatif","0.30 × r_C1","0.25 × r_C2","0.20 × r_C3","0.15 × r_C4","0.10 × r_C5","V_i"]
ws_manual.merge_cells("H28:H28")
for col, v in enumerate(vi_hdr, 2):
    cell = ws_manual.cell(28, col, v)
    cell.font  = hdr_font(10)
    cell.fill  = fill(C_HEADER_MID)
    cell.alignment = center()
    cell.border = border()

vi_results = []
for i, alt in enumerate(sample_5):
    row = 29 + i
    no, name, brand, tkdn, ram, ssd, warranty, price = alt
    r1 = round(tkdn / max_tkdn, 4)
    r2 = round(ram  / max_ram,  4)
    r3 = round(ssd  / max_ssd,  4)
    r4 = round(warranty / max_war, 4)
    r5 = round(min_price / price,  4)
    w_r1 = round(0.30*r1, 4)
    w_r2 = round(0.25*r2, 4)
    w_r3 = round(0.20*r3, 4)
    w_r4 = round(0.15*r4, 4)
    w_r5 = round(0.10*r5, 4)
    vi   = round(w_r1+w_r2+w_r3+w_r4+w_r5, 4)
    vi_results.append((no, vi))
    vals = [f"A{no:02d}", w_r1, w_r2, w_r3, w_r4, w_r5, vi]
    for col, v in enumerate(vals, 2):
        cell = ws_manual.cell(row, col, v)
        cell.font  = body_font(bold=(col==8))
        cell.fill  = fill(C_ROW_ALT) if i % 2 == 0 else fill(C_WHITE)
        cell.border = border("thin")
        cell.alignment = center()
        cell.number_format = "0.0000"

ws_manual.merge_cells("B35:G35")
s5 = ws_manual.cell(35, 2, "LANGKAH 5: Ranking – Alternatif dengan V_i TERBESAR mendapat RANKING TERTINGGI")
s5.font = Font(name="Arial", size=11, bold=True, color=C_WHITE)
s5.fill = fill(C_HEADER_MID)
s5.alignment = center()

vi_sorted = sorted(vi_results, key=lambda x: x[1], reverse=True)
rk_hdr = ["Ranking","Kode Alternatif","Nilai V_i","Keterangan"]
for col, v in enumerate(rk_hdr, 2):
    cell = ws_manual.cell(36, col, v)
    cell.font  = hdr_font(10)
    cell.fill  = fill(C_HEADER_MID)
    cell.alignment = center()
    cell.border = border()

for i, (no, vi) in enumerate(vi_sorted):
    row = 37 + i
    ket = "TERPILIH" if i == 0 else ""
    vals = [i+1, f"A{no:02d}", vi, ket]
    for col, v in enumerate(vals, 2):
        cell = ws_manual.cell(row, col, v)
        cell.font  = Font(name="Arial", size=10, bold=(i==0))
        cell.fill  = fill(C_GREEN_LIGHT) if i == 0 else (fill(C_ROW_ALT) if i%2==0 else fill(C_WHITE))
        cell.border = border("thin")
        cell.alignment = center()
        if col == 4:
            cell.number_format = "0.0000"

# ── freeze panes on key sheets ───────────────────────────────────────────────
for ws in [ws_data, ws_matrix, ws_fuzzy, ws_norm, ws_rank, ws_hasil]:
    ws.freeze_panes = ws.cell(5, 4) if ws != ws_hasil else ws.cell(5, 5)

# ── sheet tab colours ────────────────────────────────────────────────────────
tab_colors = {
    "COVER":                      "1F4E79",
    "DATA ALTERNATIF":            "2E75B6",
    "BOBOT & KRITERIA":           "F4B942",
    "MATRIKS KEPUTUSAN":          "2E75B6",
    "NILAI FUZZY":                "70AD47",
    "NORMALISASI":                "70AD47",
    "NILAI PREFERENSI & RANKING": "70AD47",
    "HASIL AKHIR":                "C00000",
    "CONTOH PERHITUNGAN":         "F4B942",
}
for name, color in tab_colors.items():
    if name in wb.sheetnames:
        wb[name].sheet_properties.tabColor = color

# ── save ─────────────────────────────────────────────────────────────────────
# Simpan di folder yang sama dengan lokasi script ini
script_dir = os.path.dirname(os.path.abspath(__file__))
out = os.path.join(script_dir, "SAW_Laptop_Kepmendag2060_2025.xlsx")
wb.save(out)
print("File Excel berhasil disimpan di:", out)